
import glob
import gpxpy
import gpxpy.gpx
import pandas as pd
import os
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime

def parse_gpx(file_path):
    with open(file_path, 'r') as gpx_file:
        gpx = gpxpy.parse(gpx_file)
        
        for track in gpx.tracks:
            for segment in track.segments:
                competitor_id = os.path.splitext(os.path.basename(file_path))[0]
                category = os.path.basename(os.path.dirname(file_path))
                if segment.points[0].time is None:
                    start_time = datetime.now()
                    finish_time = datetime.now()
                else:
                    start_time = segment.points[0].time.replace(tzinfo=None)
                    finish_time = segment.points[-1].time.replace(tzinfo=None)
                distance2d = segment.length_2d()/1000  # Distance in kilometers
                distance3d = segment.length_3d()/1000  # Distance in kilometers
                elapsed_time = (finish_time - start_time).total_seconds()
                no_time = False
                if elapsed_time < 1: 
                    no_time = True
                
                yield competitor_id, category, start_time, finish_time, elapsed_time, distance2d, distance3d, no_time

timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f'/Users/pamui/Documents/Developer/PythonPrograms/collect_gpx_data/results/race_results_{timestamp_str}.xlsx'
def write_to_excel(data, output_file=output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Race_Data"

    # Headers
    ws.append(['Competitor Id', 'Category', 'Start Time', 'Finish Time', 'Total time', 'Total Distance 2d', 'Total Distance 3d', 'No Time'])

    # Total time column
    total_time_column = None
    for cell in ws[1]:
        if cell.value == 'Total time':
            total_time_column = cell.column_letter
            break

    # Data
    for row in data:
        ws.append(row)

    # Formatting duration as [h]:mm:ss
    duration_style = NamedStyle(name='duration', number_format='[h]:mm:ss')
    for cell in ws[total_time_column][1:]:  # Column E contains durations in seconds
        # Convert seconds to Excel time format (fraction of 24 hours)
        cell.value = cell.value / 86400  # There are 86400 seconds in a day
        cell.style = duration_style

    # Define the full range of your table
    max_row = ws.max_row
    max_col = ws.max_column
    end_cell = f'{get_column_letter(max_col)}{max_row}'
    table_range = f"A1:{end_cell}"

    # Create the table
    table = Table(displayName="RaceResults", ref=table_range)

    # Add a default table style (optional)
    table_style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=True,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = table_style

    ws.add_table(table)

    wb.save(output_file)

# Path where your GPX files are stored
gpx_files = glob.glob(f'/Users/pamui/Documents/Developer/PythonPrograms/collect_gpx_data/assets/**/*.gpx', recursive=True)

# Collect data from each file
all_data = []
for file_path in tqdm(gpx_files, desc= 'Processing GPX Files'):
    all_data.extend(parse_gpx(file_path))

# Write the collected data to an Excel file
write_to_excel(all_data)
