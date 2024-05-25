
import glob
import json
import gpxpy
import gpxpy.gpx
import pandas as pd
import os
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime

with open('config.json', 'r') as config_file:
    config = json.load(config_file)

input_folder = config['folders']['input_files_folder']
reference_distance_folder = config['folders']['reference_distance_folder']
competitors_file = config['folders']["competitors_db"]['competitors_file']
competitors_sheet = config['folders']["competitors_db"]['competitors_sheet']
competitors_list = config['folders']["competitors_db"]['competitors_table']
results_output_folder = config['folders']['output_file_folder']
results_table_name = config['excel_params']['table_name']
results_sheet_name = config['excel_params']['sheet_name']
results_table_style_name = config['excel_params']['table_style_name']

def get_competitor_data(competitor_id):
    row = competitors_data[competitors_data['competitor_id'] == competitor_id]
    if not row.empty:
        team = row.iloc[0]['team']
        name = row.iloc[0]['name']
        surname = row.iloc[0]['surname']
        bike_brand = row.iloc[0]['bike_brand']
        bike_model = row.iloc[0]['bike_model']
        category = row.iloc[0]['category']
        gas_compensation = 0
        chicken_way_penalty = 0
        if row.iloc[0]['gas_compensation'] == True: gas_compensation = 6
        if row.iloc[0]['chicken_way_penalty'] == True: chicken_way_penalty = 1
        return team, name, surname, bike_brand, bike_model, category, gas_compensation, chicken_way_penalty
    return None, None, None, None, None, None, None, None

def get_reference_data(category):
    row = reference_data[reference_data['category'] == category]
    if not row.empty:
        return row.iloc[0]['min_distance3d'] 
    return 0

def parse_reference_gpx(file_path):
    with open(file_path, 'r') as gpx_file:
        gpx = gpxpy.parse(gpx_file)
        
        for track in gpx.tracks:
            for segment in track.segments:
                category = os.path.splitext(os.path.basename(file_path))[0]
                min_distance3d = segment.length_3d()/1000  # Distance in kilometers
                
                yield category, min_distance3d

def parse_gpx(file_path):
    with open(file_path, 'r') as gpx_file:
        gpx = gpxpy.parse(gpx_file)
        
        for track in gpx.tracks:
            for segment in track.segments:
                competitor_id = os.path.splitext(os.path.basename(file_path))[0]
                team, name, surname, bike_brand, bike_model, category, gas_compensation, chicken_way_penalty = get_competitor_data(competitor_id)
                if segment.points[0].time is None:
                    start_time = datetime.now()
                    finish_time = datetime.now()
                else:
                    start_time = segment.points[0].time.replace(tzinfo=None)
                    finish_time = segment.points[-1].time.replace(tzinfo=None)
                distance3d = segment.length_3d()/1000  # Distance in kilometers
                adjusted_distance = distance3d - gas_compensation + chicken_way_penalty
                min_distance3d = get_reference_data(category)
                distance_ok = True
                distance_diference = adjusted_distance - min_distance3d
                if adjusted_distance < min_distance3d: distance_ok = False
                elapsed_time = (finish_time - start_time).total_seconds()
                no_time = False
                if elapsed_time < 1: 
                    no_time = True
                
                yield competitor_id, team, name, surname, bike_brand, bike_model, category, start_time, finish_time, elapsed_time, distance3d, min_distance3d, gas_compensation, chicken_way_penalty, adjusted_distance, distance_diference, distance_ok, no_time

timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f'{results_output_folder}/race_results_{timestamp_str}.xlsx'
def write_to_excel(data, output_file=output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = results_sheet_name

    # Headers
    ws.append(['Competitor Id', 'Team', 'Name', 'Surname', 'Bike Brand', 'Bike Model', 'Category', 'Start Time', 'Finish Time', 'Total time', 'Total Distance 3d', 'Min Distance 3d', 'Gas Compensation', 'Chicken Way Penalty', 'Adjusted Total Distance', 'Distance Difference', 'Distance Ok', 'No Time'])

    # Total time column
    total_time_column = None
    for cell in ws[1]:
        if cell.value == 'Total time':
            total_time_column = cell.column_letter
            break

    # Data
    for index, row in data.iterrows():
        ws.append(row.values.tolist())

    # Formatting duration as [h]:mm:ss
    duration_style = NamedStyle(name='duration', number_format='[h]:mm:ss')
    for cell in ws[total_time_column][1:]:  # Column E contains durations in seconds
        # Convert seconds to Excel time format (fraction of 24 hours)
        cell.value = cell.value / 86400  # There are 86400 seconds in a day
        cell.style = duration_style

    # Define the full range of your table
    max_row = ws.max_row
    max_col = ws.max_column
    end_cell = f'{get_column_letter(max_col)}{max_row}'
    table_range = f"A1:{end_cell}"

    # Create the table
    table = Table(displayName=results_table_name, ref=table_range)

    # Add a default table style (optional)
    table_style = TableStyleInfo(name=results_table_style_name, showFirstColumn=True,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = table_style

    ws.add_table(table)

    wb.save(output_file)

print('Running script')
# Collect data from competitors file
print('Processing data from competitors file')
competitors_data = pd.read_excel(competitors_file, sheet_name=competitors_sheet, usecols='A:I', dtype={'category': str})
competitors_data = pd.DataFrame(competitors_data)
print('Competitors data processed')

# Path where reference GPX files are stored
reference_gpx_files = glob.glob(f'{reference_distance_folder}/**/*.gpx', recursive=True)

# Collect data from each reference file
reference_data = []

for file_path in tqdm(reference_gpx_files, desc= 'Processing reference GPX Files'):
    reference_data.extend(parse_reference_gpx(file_path))

reference_data = pd.DataFrame(reference_data, columns=['category', 'min_distance3d'])
print('Reference GPX files processed')

# Path where your GPX files are stored
gpx_files = glob.glob(f'{input_folder}/**/*.gpx', recursive=True)

# Collect data from each file
all_data = []
for file_path in tqdm(gpx_files, desc= 'Processing competitors GPX Files'):
    all_data.extend(parse_gpx(file_path))

all_data = pd.DataFrame(all_data, columns=['competitor_id', 'team', 'name', 'surname', 'bike_brand', 'bike_model', 'category', 'start_time', 'finish_time', 'elapsed_time', 'distance3d', 'min_distance3d', 'gas_compensation', 'chicken_way_penalty', 'adjusted_distance', 'distance_diference', 'distance_ok', 'no_time'])
print('Competitors GPX files processed')
# Write the collected data to an Excel file
print('Writting data to Excel')
write_to_excel(all_data)
print('Data writen to Excel')
print('Success! End of script')